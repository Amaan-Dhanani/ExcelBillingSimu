from quart import Quart, websocket  # type: ignore
import openpyxl  # type: ignore
from openpyxl.styles.fills import PatternFill  # type: ignore
from openpyxl.styles import Border, Side, Alignment # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from datetime import datetime

# SoftMaker PlanMaker workaround: ignore 'extLst' arg in PatternFill
_original_init = PatternFill.__init__
PatternFill.__init__ = lambda self, *a, **k: _original_init(
    self, *a, **{x: y for x, y in k.items() if x != "extLst"}
)

app = Quart(__name__)


async def handle_excel_error(e):
    """Centralized Excel error handling."""
    msg = str(e)
    if "extLst" in msg:
        return None  # Ignore SoftMaker error
    if isinstance(e, PermissionError):
        return {"status": "error", "message": "The Excel file is open. Close it before editing."}
    return {"status": "error", "message": msg}


@app.websocket("/ws")
async def ws_handler():
    while True:
        data = await websocket.receive_json()
        op = data.get("operation")
        file_path = data.get("data", {}).get("file_path")
        sheet = data.get("data", {}).get("sheet")
        submitted_data = data.get("data", {}).get("submitted_data", [])
        read_data = data.get("data", {}).get("read_data", [])
        row_count = None
        col_count = 0
        row_index = 19
        col_index = 2  # B = 2
        override_rows = 4

        try:
            if op == "edit_excel":
                wb = openpyxl.load_workbook(file_path)
                ws = wb[sheet]
                def unmerge_cell(cell_address):
                    for merged_range in ws.merged_cells.ranges:
                        if cell_address in merged_range:
                            ws.unmerge_cells(str(merged_range))
                            print(f"Unmerged {merged_range} containing {cell_address}")
                            break


                cell_change_lengths = [
                    "A1", "A2", "B3", "B4", "B5", "B6", "B7", "B8", "B9",
                    "A10", "B11", "B12", "B13", "B14", "A15", "B16", "B17", "B18"
                ]

                for cell in cell_change_lengths:
                    unmerge_cell(cell)

                # deletes col B
                ws.delete_cols(col_index + 1, ws.max_column - col_index) # type: ignore

                row_count = len(read_data)
                ws.delete_rows(23, row_count - 4) # type: ignore

                extra_rows_needed = len(read_data) - override_rows
                if extra_rows_needed > 0:
                    ws.insert_rows(row_index + override_rows, amount=extra_rows_needed)

                # Fill the data with outlines
                for r_idx, row in enumerate(submitted_data):
                    for c_idx, value in enumerate(row):
                        c = ws.cell(row=row_index + r_idx, column=col_index + c_idx, value=value)
                        c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) # type: ignore



                def merge_cell(cell_address):
                    # Extract the row number from the cell address (works with multi-digit rows)
                    row_count = len(read_data)
                    temp= row_count + 1
                    import re
                    match = re.match(r"([A-Z]+)(\d+)", cell_address)
                    if not match:
                        raise ValueError(f"Invalid cell address: {cell_address}")
                    
                    col_letters, row_number = match.groups()
                    
                    # Build merge range from start cell to last column
                    final_str = f"{cell_address}:{get_column_letter(temp)}{row_number}"
                    
                    # Merge cells in worksheet
                    ws.merge_cells(final_str)
                    return final_str
                
                for cell in cell_change_lengths:
                    merge_cell(cell)

                
                col_letter = get_column_letter(row_count + 2)


                def vert_line(col_letter):
                    for r in range(1, ws.max_row + 1):
                        first_cell_value = ws.cell(row=r, column=1).value
                        ws[f"{col_letter}{r}"].border = Border(left=Side(style='medium'))
                        if first_cell_value == "NDC code (If applicable)":
                            break
                vert_line(get_column_letter(row_count + 2))
                vert_line("A")

                for r in range(1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=1)  # Column A
                    
                    # Add thin bottom border, preserving existing sides
                    cell.border = Border(
                        left=cell.border.left if cell.border else Side(style=None),
                        right=cell.border.right if cell.border else Side(style=None),
                        top=cell.border.top if cell.border else Side(style=None),
                        bottom=Side(style='thin')
                    )

                    # Check for "Date(s) of Service"
                    if cell.value == "Date(s) of Service":
                        merge_start = r
                        merge_end = r
                        while merge_end + 1 <= ws.max_row and ws.cell(row=merge_end + 1, column=1).value in (None, ""):
                            merge_end += 1
                        
                        if merge_end > merge_start:
                            col_letter = get_column_letter(1)  # "A"
                            ws.merge_cells(f"{col_letter}{merge_start}:{col_letter}{merge_end}")
                            merged_cell = ws.cell(row=merge_start, column=1)
                            # Vertical center only, keep horizontal as is (usually left)
                            merged_cell.alignment = Alignment(vertical="center")

                    # Stop after "NDC code (If applicable)"
                    if cell.value == "NDC code (If applicable)":
                        break

                def merge_second_cell_row(search_str, end_col):
                    for row in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=1).value
                        if cell_value == search_str:
                            start_col = 2  # column B
                            start_cell = ws.cell(row=row, column=start_col).coordinate
                            end_cell_letter = get_column_letter(end_col)
                            end_cell = f"{end_cell_letter}{row}"
                            ws.merge_cells(f"{start_cell}:{end_cell}")
                            return
                    print(f"'{search_str}' not found in column A.")
                
                merge_second_cell_row("Facility information (Name/NPI)", row_count+1)
                merge_second_cell_row("NDC code (If applicable)", row_count+1)
                
                wb.save(file_path)

                await websocket.send_json({
                    "status": "edit_made",
                    "message": "hello",
                })
                

            elif op == "get_sheets":
                wb = openpyxl.load_workbook(file_path)
                await websocket.send_json(
                    {"status": "received_sheets", "message": wb.sheetnames}
                )

            elif op == "each_sheet":
                wb = openpyxl.load_workbook(file_path)
                ws = wb[sheet]

                # Count consecutive non-blank cells in row 19 starting at B
                for _ in range(1000000):  # practically infinite loop
                    col_letter = get_column_letter(col_index)
                    cell_value = ws[f"{col_letter}{row_index}"].value

                    if cell_value is None or str(cell_value).strip() == "":
                        break  # stop when blank

                    col_count += 1
                    col_index += 1  # move right

                # Find row count by locating "Facility information (Name/NPI)"
                target = "Facility information (Name/NPI)"

                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value == target:
                            row_number = cell.row
                            row_count = row_number - 19
                            break
                    if row_count is not None:
                        break

                if row_count is None:
                    print(f"'{target}' not found")
                    row_count = 0
                else:
                    print(f"Row difference: {row_count}")

                print("Consecutive non-blank cells:", col_count)

                # Build 2D array from B19 right col_count and down row_count
                table_data = []
                for r in range(row_index, row_index + row_count):
                    row_data = []
                    for c in range(2, 2 + col_count):
                        cell = ws.cell(row=r, column=c)
                        raw_val = getattr(cell, "_value", cell.value)
                        
                        def convert_if_needed(raw_val):
                            conv_str = str(raw_val)
                            try:
                                dt = datetime.strptime(conv_str, "%Y-%m-%d %H:%M:%S")
                                return dt.strftime("%m/%d/%Y")
                            except ValueError:
                                 return "" if conv_str == "None" else conv_str

                        # When appending:
                        row_data.append(convert_if_needed(raw_val))
                    table_data.append(row_data)

                # Send result
                await websocket.send_json(
                    {
                        "status": "table_data",
                        "message": table_data,
                    }
                )

        except Exception as e:
            err = await handle_excel_error(e)
            if err:
                await websocket.send_json(err)


if __name__ == "__main__":
    app.run(port=5000, host="0.0.0.0")